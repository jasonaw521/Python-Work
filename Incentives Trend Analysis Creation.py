#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import sys
import os
from os import path
from datetime import date
import numpy as np
import xlwings as xl
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import shutil
import warnings
warnings.simplefilter("ignore")
import win32com.client as w3c
pd.io.formats.excel.ExcelFormatter.header_style = None


# In[2]:


###User Information
user = os.getlogin()
current_month = date.today().strftime('%Y-%m')
if user == 'span':
    template_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Incentives\\scripts\\Incentives Review Automation Template\\Incentives Review Template.xlsx')
else:
    template_path =  os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\\Incentives\\scripts\\Incentives Review Automation Template\\Incentives Review Template.xlsx')
if os.path.exists(template_path):
    print(template_path)
else:
    print(f'{template_path} does not exist!')


# In[3]:


###Find Folders
if user == 'span':
    it_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Incentives',current_month)
else:
    it_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\\Incentives',current_month)

if os.path.exists(it_path):
    print(it_path)
else:
    print(f'{it_path} does not exist.')
    current_month = input('Input new current_month.')
    it_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\Incentives',current_month)


# In[5]:


###This cell creates a new trend file for EVERY OEM (not BMW or MINI)

skipped_oems = []
for dirnames in os.listdir(it_path):
    if not "BMW" in dirnames:
        if not "MINI" in dirnames:
            if not "_" in dirnames:
                if os.path.isdir(os.path.join(it_path,dirnames)):
                    print(os.path.join(it_path,dirnames))
                    oem_path = os.path.join(it_path,dirnames)
                    oem = os.path.split(oem_path)[1]
                    dfc = pd.DataFrame()
                    dfdt = pd.DataFrame()
                    dfd = pd.DataFrame()
                    dfp = pd.DataFrame()
                    for file in os.listdir(oem_path):
                        #Per File
                        if file.endswith(".csv"):
                            if not "Replacement" in file:
                                if "brandcashincentives" in file:
                                    dfc = pd.read_csv(os.path.join(oem_path,file))
                                if "CompatibilityByDealType" in file:
                                    dfdt = pd.read_csv(os.path.join(oem_path,file))
                                if "CompatibilityByDescription" in file:
                                    dfd = pd.read_csv(os.path.join(oem_path,file))
                                if "programPresence" in file:
                                    dfp = pd.read_csv(os.path.join(oem_path,file))

                    #Per OEM
                    
                    file = template_path
                    writer = pd.ExcelWriter(file, engine = 'openpyxl',mode ='a',if_sheet_exists = 'replace') 
                    dfc.to_excel(writer,sheet_name = 'CashIncentivesReport',index=False)
                    dfdt.to_excel(writer,sheet_name = 'DealTypeCompatibility',index=False)
                    dfd.to_excel(writer,sheet_name = 'DescriptionCompatibility',index=False)
                    dfp.to_excel(writer,sheet_name = 'ProgramPresence',index=False)
                    writer.save()
                    writer.close()
                    
                    book = load_workbook(file)
                    book.save(os.path.join(oem_path,f'Incentives Trend Analysis {oem}-{current_month}.xlsx'))
                    book.close()
        

                    print(f'Trend analysis file saved for {oem}.')
                
input(f'The following brands were skipped due to no trend analysis files : {skipped_oems}. Press enter to continue.')


# In[ ]:


###Use this cell to only create a Trend Analysis File for a Specific OEM.

while True:
    oem = input('Please enter the name of the OEM you want to check ...')
    #current_month = input('Please enter current month.')
    spath = os.path.join('C:\\Users',user,'automotivemastermind','Data Ingestion - Documents','Incentives',current_month,oem)
    if path.exists(spath):
        print(spath)
        break
    else:
        print(spath)
        print('Invalid input. Try again.')

dfc1 = pd.DataFrame()
dfdt1 = pd.DataFrame()
dfd1 = pd.DataFrame()
dfp1 = pd.DataFrame()

for file in os.listdir(spath):
    if file.endswith(".csv"):
        if not "Replacement" in file:
            if "brandcashincentives" in file:
                dfc1 = pd.read_csv(os.path.join(spath,file))
            if "CompatibilityByDealType" in file:
                dfdt1 = pd.read_csv(os.path.join(spath,file))
            if "CompatibilityByDescription" in file:
                dfd1 = pd.read_csv(os.path.join(spath,file))
            if "programPresence" in file:
                dfp1 = pd.read_csv(os.path.join(spath,file))
            else:
                pass 
            
file = template_path
writer = pd.ExcelWriter(file, engine = 'openpyxl',mode ='a',if_sheet_exists = 'replace')#,engine_kwargs={'data_only':True})

if dfc1.empty == False:
    dfc1.to_excel(writer,sheet_name = 'CashIncentivesReport',index=False)
    print(f'Cash Incentives file loaded for {oem}.')
else:
    print(f'Cash Incentives file not found for {oem}.')
    
if dfdt1.empty == False:
    dfdt1.to_excel(writer,sheet_name = 'DealTypeCompatibility',index=False)
    print(f'CompatibilityByDealType file loaded for {oem}')
else:
    print(f'CompatibilityByDealType file not found for {oem}.')
if dfd1.empty == False:
    dfd1.to_excel(writer,sheet_name = 'DescriptionCompatibility',index=False)
    print(f'CompatibilityByDescription file loaded for {oem}.')
else:
    print(f'CompatibilityByDescription file not found for {oem}.')
if dfp1.empty == False:
    dfp1.to_excel(writer,sheet_name = 'ProgramPresence',index=False)
    print(f'ProgramPresence file loaded for {oem}.')
else:
    print(f'programPresence file not found for {oem}.')
    
writer.save()
writer.close()

if dfd1.empty == False:
    book = load_workbook(file)
    book.save(os.path.join(spath,f'Incentives Trend Analysis {oem}-{current_month}.xlsx'))
    book.close()
else:
    pass

print(f'Trend analysis file saved for {oem} at {spath}.')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




