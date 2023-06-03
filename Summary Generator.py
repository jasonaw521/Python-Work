#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import sys
import os
from os import path
from datetime import date
import numpy as np
import xlwings as xw
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import shutil
import warnings
warnings.simplefilter("ignore")
import win32com.client as w3c
from datetime import date
current_date = date.today().strftime('%Y-%m-%d')
pd.io.formats.excel.ExcelFormatter.header_style = None
from win32com.client import Dispatch
import time
from openpyxl.utils import get_column_letter


# In[2]:


def ObtainUserInfo():
    user = os.getlogin()
    global current_month
    global current_day
    current_day = date.today()
    current_month = date.today().strftime('%Y-%m')
    month = date.today().strftime('%m')
    year = date.today().strftime('%Y')
    smallyear = date.today().strftime('%y')
    global save_path
    if os.path.exists(os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\\Incentives\\FAQs\\summaries',year,month)):
        save_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\\Incentives\\FAQs\\summaries',year,month)
    else:
        if os.path.exists(os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Incentives\\FAQs\\summaries',year,month)):
            save_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Incentives\\FAQs\\summaries',year,month)
        else:
            save_path = input('Paste the path of the summaries folder.')
    if os.path.exists(save_path):
        input(f'Confirm Save Path : {save_path}')
    else:
        os.mkdir(save_path)
        input(f'Confirm Save Path : {save_path}')
    global mappings_path
    if os.path.exists(os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\\Incentives\\FAQs\\summaries',year,month,f'{smallyear}_{month}_incentives_summary.xlsx')):
        mappings_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Documents\\Incentives\\FAQs\\summaries',year,month,f'{smallyear}_{month}_incentives_summary.xlsx')
    else:
        if os.path.exists(os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Incentives\\FAQs\\summaries',year,month,f'{smallyear}_{month}_incentives_summary.xlsx')):
            mappings_path = os.path.join('C:\\Users',user,'automotiveMastermind\\Data Ingestion - Incentives\\FAQs\\summaries',year,month,f'{smallyear}_{month}_incentives_summary.xlsx')
        else:
            mappings_path = input('Input filepath of the monthly incentives summary.')
                
def move_sheet(wb, from_loc=None, to_loc=None):
    sheets=wb._sheets
    if from_loc is None:
        from_loc = len(sheets) - 1
    if to_loc is None:
        to_loc = 0
    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)


# In[3]:


def get_col_widths(dataframe):
    # First we find the maximum length of the index column   
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

    
def read_main():
    file_input = input(f'Paste file_path of All_Active_Programs file.')
    mainfile = file_input[1:-1]
    global main_df
    global cyient_df
    global field_df
    main_df = pd.read_csv(mainfile)
    cyient_df = main_df.copy()
    field_df = main_df.copy()
    
def write_field():
    in_use_dropped_df = field_df[field_df['In Use']==True]
    field_dropped_df = in_use_dropped_df.drop(['Program Description Original','In Use','Eligibility Original','Deal Type','Deal Type Original','Program Compatibility','Program Dates','Term for Rates','Overrides','Models','Eligibility Requirements','Previous Ownership','Group Affiliation', 'Previous Ownership Detail','Group Affiliation Detail' ], axis=1)
    writer = pd.ExcelWriter(os.path.join(save_path,f"{current_date}_Incentive_Mappings.xlsx"), engine='xlsxwriter')
    field_dropped_df.to_excel(writer,index=False,sheet_name='IncentivesInfo')
    workbook  = writer.book
    worksheet = writer.sheets['IncentivesInfo']
    widths = get_col_widths(field_dropped_df)
    format1 = workbook.add_format({'bold': True,'align':'center'})
    format2 = workbook.add_format({'bg_color':'#548235','font_color':'white','border':1})
    format3 = workbook.add_format({'bg_color':'#3A0074','font_color':'white','bold':True,'align':'center'})
    worksheet.set_row(0, None, format1)
    worksheet.conditional_format('B1:Z1', {'type':'no_blanks','format': format2})
    worksheet.write_string(0,0,'Brands',format3)
    for i, width in enumerate(get_col_widths(field_dropped_df.iloc[:,:-1])):
        worksheet.set_column(i-1, i-1, width+5)
    worksheet.autofilter(0,0,5,5)
    worksheet.freeze_panes('B2')
    worksheet.set_column('F:F',15)
    workbook.close()
    writer.close()
    
    time.sleep(3)
    
    #work1 = load_workbook(os.path.join(save_path,f"{current_date}_Incentive_Mappings.xlsx"))
    #work1.create_sheet('Mapping')
    #work1.save(os.path.join(save_path,f"{current_date}_Incentive_Mappings.xlsx"))
    
    global mappings_path

    
    dest_path = os.path.join(save_path,f"{current_date}_Incentive_Mappings.xlsx")
    with xw.App(visible=False) as app:                   
        if os.path.exists(mappings_path):
            map_book = xw.Book(mappings_path)
            dest_book = xw.Book(dest_path)
        else:
            mappings_path = input('Please paste the path of the monthly incentives summary that includes the mappings.')
            map_book = xw.Book(mappings_path)
            dest_book = xw.Book(dest_path)

        map_sheet = map_book.sheets("MAPPINGS")
        dest_sheet = dest_book.sheets.add("MAPPINGS")
        map_sheet.range('A1:J99').copy(dest_sheet.range('A1:J99'))
        map_book.close()
        dest_book.save()
        dest_book.close()
        
    time.sleep(3)

    workbook = load_workbook(os.path.join(save_path,f"{current_date}_Incentive_Mappings.xlsx"))
    move_sheet(workbook,1,0)
    worksheet = workbook["MAPPINGS"]
    worksheet['A1'].fill = PatternFill("solid", fgColor = "3A0074")
    worksheet['B1'].fill = PatternFill("solid", fgColor = "548235")
    worksheet['E1'].fill = PatternFill("solid", fgColor = "548235")
    worksheet['H1'].fill = PatternFill("solid", fgColor = "548235")
    worksheet.column_dimensions['A'].width = 30
    colu_list = ['B','C','D','E','F','G','H','I','J']
    for colu in colu_list:
        worksheet.column_dimensions[colu].width = 25
    workbook.active = workbook['IncentivesInfo']
    workbook["MAPPINGS"].views.sheetView[0].tabSelected = False
    workbook.save(os.path.join(save_path,f"{current_date}_Incentive_Mappings.xlsx"))
    workbook.close
    
def write_cyient():
    cyient_dropped_df = cyient_df.drop(['Program Description Original','Eligibility Original','Deal Type','Deal Type Original','Program Compatibility','Program Dates','Term for Rates','Overrides','Models','Eligibility Requirements','Previous Ownership','Group Affiliation', 'Previous Ownership Detail','Group Affiliation Detail' ], axis=1)
    writer = pd.ExcelWriter(os.path.join(save_path,f"{current_date}_Cyient_Summary.xlsx"), engine='xlsxwriter')
    cyient_dropped_df.to_excel(writer,index=False,sheet_name='IncentivesInfo')
    workbook  = writer.book
    worksheet = writer.sheets['IncentivesInfo']
    widths = get_col_widths(cyient_dropped_df)
    format1 = workbook.add_format({'bold': True,'align':'center'})
    format2 = workbook.add_format({'bg_color':'#548235','font_color':'white','border':1})
    format3 = workbook.add_format({'bg_color':'#3A0074','font_color':'white','bold':True,'align':'center'})
    worksheet.set_row(0, None, format1)
    worksheet.conditional_format('B1:Z1', {'type':'no_blanks','format': format2})
    worksheet.write_string(0,0,'Brands',format3)
    for i, width in enumerate(get_col_widths(cyient_dropped_df.iloc[:,:-1])):
        worksheet.set_column(i-1, i-1, width+5)
    worksheet.autofilter(0,0,5,5)
    worksheet.freeze_panes('B2')
    worksheet.set_column('G:G',15)
    workbook.close()
    writer.close()


# In[4]:


ObtainUserInfo()
read_main()
write_field()
time.sleep(3)
write_cyient()
print("Summaries Generated!")


# In[ ]:


print(mappings_path)


# In[ ]:


os.path.exists(mappings_path)
map_book = xw.Book(mappings_path)
map_sheet = map_book.sheets("MAPPINGS").used_range.value
df = pd.DataFrame(map_sheet)
print(df)


# In[ ]:




